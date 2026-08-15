from __future__ import annotations

import hashlib
import io
import json
import os
import tempfile
import unittest
import zipfile
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook

from tests.task9_fixtures import invented_aligned, invented_grouping
from tritrack_editing_assistant import organizer, paper_edit
from tritrack_editing_assistant.contracts import validate_contract


def write_aligned(root: Path, *, formula_text: bool = False) -> tuple[Path, bytes]:
    root.mkdir(parents=True, exist_ok=True)
    aligned = invented_aligned()
    if formula_text:
        aligned["takes"][0]["cues"][0]["text"] = "=INVENTED()"
    encoded = (
        json.dumps(aligned, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    ).encode("utf-8")
    path = root / "aligned.json"
    path.write_bytes(encoded)
    return path, encoded


def write_grouping(root: Path, aligned_bytes: bytes) -> tuple[Path, bytes]:
    grouping = invented_grouping()
    grouping["alignedTranscriptSha256"] = hashlib.sha256(aligned_bytes).hexdigest()
    encoded = (
        json.dumps(grouping, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    ).encode("utf-8")
    path = root / "grouping.json"
    path.write_bytes(encoded)
    return path, encoded


def logical_grid(path: Path) -> dict[str, list[list[object]]]:
    workbook = load_workbook(path, data_only=False)
    return {
        name: [list(row) for row in workbook[name].iter_rows(values_only=True)]
        for name in workbook.sheetnames
    }


class PaperExportTest(unittest.TestCase):
    def test_exports_exact_reference_grid_and_hidden_manifest(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            aligned, aligned_bytes = write_aligned(root)
            output = root / "paper.xlsx"

            summary = paper_edit.export_workbook(
                aligned,
                grouping_path=None,
                output_path=output,
            )

            workbook = load_workbook(output, data_only=False)
            self.assertEqual(
                workbook.sheetnames,
                ["Cues", "Questions", "Selections", "_TriTrack"],
            )
            self.assertEqual(workbook["_TriTrack"].sheet_state, "hidden")
            self.assertEqual(
                [cell.value for cell in workbook["Cues"][1]],
                [
                    "TakeId",
                    "SourceSha256",
                    "CueId",
                    "StartMs",
                    "EndMs",
                    "Text",
                    "Disposition",
                ],
            )
            self.assertEqual(workbook["Cues"].max_row, 5)
            self.assertEqual(
                [cell.value for cell in workbook["Cues"][2]],
                [
                    "A.wav",
                    "3" * 64,
                    "cue-000001",
                    0,
                    500,
                    "Invented first answer.",
                    "original",
                ],
            )
            self.assertEqual(workbook["Questions"].max_row, 1)
            self.assertEqual(workbook["Selections"].max_row, 1)
            manifest = {
                workbook["_TriTrack"].cell(row=row, column=1).value:
                workbook["_TriTrack"].cell(row=row, column=2).value
                for row in range(2, workbook["_TriTrack"].max_row + 1)
            }
            self.assertEqual(
                manifest["WorkbookSchemaVersion"],
                "tritrack.paper-workbook/v1",
            )
            self.assertEqual(
                manifest["AlignedTranscriptSha256"],
                hashlib.sha256(aligned_bytes).hexdigest(),
            )
            self.assertEqual(
                summary,
                {
                    "cueCount": 4,
                    "questionCount": 0,
                    "selectionCount": 0,
                },
            )

    def test_prefills_grouping_as_a_direct_projection(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            aligned, aligned_bytes = write_aligned(root)
            grouping, _ = write_grouping(root, aligned_bytes)
            output = root / "prefilled.xlsx"

            summary = paper_edit.export_workbook(
                aligned,
                grouping_path=grouping,
                output_path=output,
            )

            workbook = load_workbook(output, data_only=False)
            self.assertEqual(
                [cell.value for cell in workbook["Questions"][2]],
                ["question-001", "What changed?", 1],
            )
            self.assertEqual(
                [cell.value for cell in workbook["Selections"][2]],
                [
                    "ANSWER",
                    "answer-001",
                    "question-001",
                    1,
                    "A.wav",
                    "cue-000001",
                    "cue-000002",
                    None,
                    "Primary invented answer",
                ],
            )
            self.assertEqual(
                [cell.value for cell in workbook["Selections"][4]],
                [
                    "RESERVE",
                    "reserve-001",
                    None,
                    1,
                    "B.wav",
                    "cue-000002",
                    "cue-000002",
                    "Alternate invented answer",
                    "Keep available",
                ],
            )
            self.assertEqual(
                summary,
                {"cueCount": 4, "questionCount": 2, "selectionCount": 3},
            )

    def test_formula_looking_transcript_text_is_saved_as_literal_string(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            aligned, _ = write_aligned(root, formula_text=True)
            output = root / "literal.xlsx"

            paper_edit.export_workbook(
                aligned,
                grouping_path=None,
                output_path=output,
            )

            workbook = load_workbook(output, data_only=False)
            cell = workbook["Cues"]["F2"]
            self.assertEqual(cell.value, "=INVENTED()")
            self.assertEqual(cell.data_type, "s")

    def test_rejects_excel_unsafe_aligned_identity_without_traceback(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            aligned_payload = invented_aligned()
            aligned_payload["takes"][0]["takeId"] = "A\v.wav"
            aligned = root / "aligned.json"
            aligned.write_text(
                json.dumps(
                    aligned_payload,
                    ensure_ascii=False,
                    indent=2,
                    sort_keys=True,
                )
                + "\n",
                encoding="utf-8",
            )
            output = root / "paper.xlsx"

            with self.assertRaisesRegex(
                ValueError,
                "TRITRACK_PAPER_ALIGNED_INVALID",
            ):
                paper_edit.export_workbook(
                    aligned,
                    grouping_path=None,
                    output_path=output,
                )
            self.assertFalse(output.exists())

    def test_existing_output_and_missing_parent_fail_closed(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            output = root / "paper.xlsx"
            output.write_text("winner", encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "TRITRACK_OUTPUT_EXISTS"):
                paper_edit.export_workbook(
                    root / "missing.json",
                    grouping_path=None,
                    output_path=output,
                )
            self.assertEqual(output.read_text(encoding="utf-8"), "winner")

            aligned, _ = write_aligned(root)
            with self.assertRaisesRegex(
                ValueError, "TRITRACK_OUTPUT_PARENT_MISSING"
            ):
                paper_edit.export_workbook(
                    aligned,
                    grouping_path=None,
                    output_path=root / "missing" / "paper.xlsx",
                )


class PaperApplyTest(unittest.TestCase):
    def editable_workbook(self, root: Path) -> tuple[Path, Path, bytes]:
        aligned, aligned_bytes = write_aligned(root)
        workbook_path = root / "editable.xlsx"
        paper_edit.export_workbook(
            aligned,
            grouping_path=None,
            output_path=workbook_path,
        )
        workbook = load_workbook(workbook_path, data_only=False)
        questions = workbook["Questions"]
        questions.append(["question-001", "  What   changed?  ", 1])
        questions.append(["question-002", "What comes next?", 2])
        selections = workbook["Selections"]
        selections.append(
            [
                "ANSWER",
                "answer-001",
                "question-001",
                1,
                "A.wav",
                "cue-000001",
                "cue-000002",
                None,
                "  Primary   invented answer  ",
            ]
        )
        selections.append(
            [
                "ANSWER",
                "answer-002",
                "question-002",
                1,
                "B.wav",
                "cue-000001",
                "cue-000001",
                None,
                None,
            ]
        )
        selections.append(
            [
                "RESERVE",
                "reserve-001",
                None,
                1,
                "B.wav",
                "cue-000002",
                "cue-000002",
                " Alternate   invented answer ",
                "  Keep   available ",
            ]
        )
        workbook.save(workbook_path)
        return aligned, workbook_path, aligned_bytes

    def test_applies_normalized_edits_to_strict_grouping(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            aligned, workbook, aligned_bytes = self.editable_workbook(root)
            output = root / "grouping.json"

            grouping = paper_edit.apply_workbook(
                aligned,
                workbook,
                output_path=output,
            )

            validate_contract("grouping-v1", grouping)
            expected = invented_grouping()
            expected["alignedTranscriptSha256"] = hashlib.sha256(
                aligned_bytes
            ).hexdigest()
            self.assertEqual(grouping, expected)
            self.assertEqual(output.read_bytes(), organizer.encode_grouping(expected))
            encoded = output.read_text(encoding="utf-8")
            self.assertNotIn("startMs", encoded)
            self.assertNotIn("sourceSha256", encoded)
            self.assertNotIn("Invented first answer.", encoded)

    def test_grouping_fixpoint_and_logical_grid_idempotence(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            aligned, aligned_bytes = write_aligned(root)
            grouping, grouping_bytes = write_grouping(root, aligned_bytes)
            workbook = root / "prefilled.xlsx"
            applied = root / "applied.json"

            paper_edit.export_workbook(
                aligned,
                grouping_path=grouping,
                output_path=workbook,
            )
            paper_edit.apply_workbook(aligned, workbook, output_path=applied)
            self.assertEqual(applied.read_bytes(), grouping_bytes)

            edited_aligned, edited_workbook, _ = self.editable_workbook(root / "edit")
            normalized = root / "normalized.json"
            paper_edit.apply_workbook(
                edited_aligned,
                edited_workbook,
                output_path=normalized,
            )
            first = root / "first.xlsx"
            second = root / "second.xlsx"
            paper_edit.export_workbook(
                edited_aligned,
                grouping_path=normalized,
                output_path=first,
            )
            paper_edit.export_workbook(
                edited_aligned,
                grouping_path=normalized,
                output_path=second,
            )
            self.assertEqual(logical_grid(first), logical_grid(second))
            first_applied = root / "first-applied.json"
            second_applied = root / "second-applied.json"
            paper_edit.apply_workbook(
                edited_aligned, first, output_path=first_applied
            )
            paper_edit.apply_workbook(
                edited_aligned, second, output_path=second_applied
            )
            self.assertEqual(first_applied.read_bytes(), normalized.read_bytes())
            self.assertEqual(second_applied.read_bytes(), normalized.read_bytes())

    def test_rejects_formula_reference_manifest_and_sheet_tampering(self) -> None:
        def reject_edit(mutator, code: str) -> None:
            with tempfile.TemporaryDirectory() as temporary:
                root = Path(temporary)
                aligned, workbook_path, _ = self.editable_workbook(root)
                workbook = load_workbook(workbook_path, data_only=False)
                mutator(workbook)
                workbook.save(workbook_path)
                with self.assertRaisesRegex(ValueError, code):
                    paper_edit.apply_workbook(
                        aligned,
                        workbook_path,
                        output_path=root / "grouping.json",
                    )

        reject_edit(
            lambda workbook: setattr(
                workbook["Questions"]["B2"], "value", "=INVENTED()"
            ),
            "TRITRACK_PAPER_FORMULA_FORBIDDEN",
        )
        reject_edit(
            lambda workbook: setattr(workbook["Cues"]["F2"], "value", "Misleading"),
            "TRITRACK_PAPER_REFERENCE_MISMATCH",
        )
        reject_edit(
            lambda workbook: setattr(
                workbook["_TriTrack"]["B4"], "value", "f" * 64
            ),
            "TRITRACK_PAPER_MANIFEST_MISMATCH",
        )
        reject_edit(
            lambda workbook: workbook.create_sheet("Unexpected"),
            "TRITRACK_PAPER_SHEETS_INVALID",
        )
        reject_edit(
            lambda workbook: workbook["Selections"].merge_cells("A2:B2"),
            "TRITRACK_PAPER_WORKBOOK_INVALID",
        )

    def test_rejects_external_cell_hyperlinks(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            aligned, workbook_path, _ = self.editable_workbook(root)
            workbook = load_workbook(workbook_path, data_only=False)
            workbook["Questions"]["B2"].hyperlink = (
                "https://example.invalid/external"
            )
            workbook.save(workbook_path)

            with self.assertRaisesRegex(
                ValueError,
                "TRITRACK_PAPER_WORKBOOK_INVALID",
            ):
                paper_edit.apply_workbook(
                    aligned,
                    workbook_path,
                    output_path=root / "grouping.json",
                )

    def test_rejects_extreme_dimensions_before_cell_iteration(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            aligned, workbook_path, _ = self.editable_workbook(root)
            workbook = load_workbook(workbook_path, data_only=False)
            workbook["Questions"].cell(
                row=1_048_576,
                column=3,
            ).number_format = "@"
            workbook.save(workbook_path)

            with (
                mock.patch(
                    "openpyxl.worksheet.worksheet.Worksheet.iter_rows",
                    side_effect=AssertionError("cell iteration started"),
                ),
                self.assertRaisesRegex(
                    ValueError,
                    "TRITRACK_PAPER_WORKBOOK_INVALID",
                ),
            ):
                paper_edit.apply_workbook(
                    aligned,
                    workbook_path,
                    output_path=root / "grouping.json",
                )

    def test_rejects_excessive_expanded_zip_before_openpyxl_load(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            workbook_path = root / "expanded.xlsx"
            buffer = io.BytesIO()
            with zipfile.ZipFile(
                buffer,
                mode="w",
                compression=zipfile.ZIP_DEFLATED,
            ) as archive:
                archive.writestr("xl/worksheets/sheet1.xml", b"x" * 4096)
            workbook_path.write_bytes(buffer.getvalue())

            with (
                mock.patch.object(
                    paper_edit,
                    "_WORKBOOK_EXPANDED_LIMIT_BYTES",
                    1024,
                ),
                mock.patch.object(
                    paper_edit,
                    "load_workbook",
                    side_effect=AssertionError("openpyxl load started"),
                ),
                self.assertRaisesRegex(
                    ValueError,
                    "TRITRACK_PAPER_WORKBOOK_INVALID",
                ),
            ):
                paper_edit._load_workbook_artifact(workbook_path)

    def test_rejects_partial_rows_bad_placement_and_foreign_spans(self) -> None:
        def reject_selection(row, code: str) -> None:
            with tempfile.TemporaryDirectory() as temporary:
                root = Path(temporary)
                aligned, _ = write_aligned(root)
                workbook_path = root / "paper.xlsx"
                paper_edit.export_workbook(
                    aligned,
                    grouping_path=None,
                    output_path=workbook_path,
                )
                workbook = load_workbook(workbook_path, data_only=False)
                workbook["Questions"].append(
                    ["question-001", "What changed?", 1]
                )
                workbook["Selections"].append(row)
                workbook.save(workbook_path)
                with self.assertRaisesRegex(ValueError, code):
                    paper_edit.apply_workbook(
                        aligned,
                        workbook_path,
                        output_path=root / "grouping.json",
                    )

        reject_selection(
            ["ANSWER", "answer-001", "question-001"],
            "TRITRACK_PAPER_ROW_INVALID",
        )
        reject_selection(
            [
                "OTHER",
                "answer-001",
                "question-001",
                1,
                "A.wav",
                "cue-000001",
                "cue-000001",
                None,
                None,
            ],
            "TRITRACK_PAPER_ROW_INVALID",
        )
        reject_selection(
            [
                "ANSWER",
                "answer-001",
                "question-001",
                1,
                "Unknown.wav",
                "cue-000001",
                "cue-000001",
                None,
                None,
            ],
            "TRITRACK_ORGANIZER_TAKE_UNKNOWN",
        )

    def test_file_boundaries_late_mutation_and_publication_race(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            aligned, workbook, _ = self.editable_workbook(root)
            symlink = root / "paper-link.xlsx"
            symlink.symlink_to(workbook)
            with self.assertRaisesRegex(ValueError, "TRITRACK_PAPER_WORKBOOK_INVALID"):
                paper_edit.apply_workbook(
                    aligned,
                    symlink,
                    output_path=root / "symlink.json",
                )

            invalid_zip = root / "invalid.xlsx"
            invalid_zip.write_bytes(b"not a workbook")
            with self.assertRaisesRegex(ValueError, "TRITRACK_PAPER_WORKBOOK_INVALID"):
                paper_edit.apply_workbook(
                    aligned,
                    invalid_zip,
                    output_path=root / "invalid.json",
                )

            output = root / "changed.json"
            original_verify = paper_edit._verify_unchanged

            def mutate_then_verify(artifact):
                if artifact.path == workbook:
                    workbook.write_bytes(workbook.read_bytes() + b" ")
                return original_verify(artifact)

            with (
                mock.patch.object(
                    paper_edit,
                    "_verify_unchanged",
                    side_effect=mutate_then_verify,
                ),
                self.assertRaisesRegex(ValueError, "TRITRACK_PAPER_INPUT_CHANGED"),
            ):
                paper_edit.apply_workbook(aligned, workbook, output_path=output)
            self.assertFalse(output.exists())

            aligned, workbook, _ = self.editable_workbook(root / "race-input")
            race_output = root / "race.json"

            def race_winner(_source, destination):
                Path(destination).write_text("winner", encoding="utf-8")
                raise FileExistsError

            with (
                mock.patch.object(os, "link", side_effect=race_winner),
                self.assertRaisesRegex(ValueError, "TRITRACK_OUTPUT_EXISTS"),
            ):
                paper_edit.apply_workbook(
                    aligned,
                    workbook,
                    output_path=race_output,
                )
            self.assertEqual(race_output.read_text(encoding="utf-8"), "winner")
            self.assertEqual(list(root.glob(".race.json.*.tmp")), [])


if __name__ == "__main__":
    unittest.main()

"""Strict XLSX transport for the cue-addressed paper-edit round trip."""

from __future__ import annotations

import hashlib
import io
import json
import os
import stat
import tempfile
import xml.etree.ElementTree as element_tree
import zipfile
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from pathlib import Path, PurePosixPath

from jsonschema import ValidationError
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils.exceptions import InvalidFileException

from . import __version__, organizer
from .contracts import validate_contract
from .process import require_absent_output

WORKBOOK_SCHEMA_VERSION = "tritrack.paper-workbook/v1"
CUES_HEADERS = (
    "TakeId",
    "SourceSha256",
    "CueId",
    "StartMs",
    "EndMs",
    "Text",
    "Disposition",
)
QUESTIONS_HEADERS = ("QuestionId", "Question", "Order")
SELECTIONS_HEADERS = (
    "Placement",
    "SegmentId",
    "QuestionId",
    "Order",
    "TakeId",
    "StartCueId",
    "EndCueId",
    "ReserveReason",
    "EditorNote",
)
MANIFEST_HEADERS = ("Key", "Value")
SHEET_NAMES = ("Cues", "Questions", "Selections", "_TriTrack")
_JSON_LIMIT_BYTES = 16 * 1024 * 1024
_WORKBOOK_LIMIT_BYTES = 64 * 1024 * 1024
_WORKBOOK_MEMBER_LIMIT = 512
_WORKBOOK_EXPANDED_LIMIT_BYTES = 256 * 1024 * 1024
_WORKBOOK_SINGLE_MEMBER_LIMIT_BYTES = 128 * 1024 * 1024


@dataclass(frozen=True)
class LoadedArtifact:
    path: Path
    payload: object
    encoded: bytes
    sha256: str
    invalid_code: str
    limit: int


@dataclass(frozen=True)
class ValidatedWorkbook:
    aligned_sha256: str
    workbook_sha256: str
    workbook_schema_version: str
    cue_count: int
    question_count: int
    answer_count: int
    reserve_count: int
    grouping: dict[str, object]


def _read_regular_bytes(path: Path, *, limit: int, invalid_code: str) -> bytes:
    flags = os.O_RDONLY
    if hasattr(os, "O_NOFOLLOW"):
        flags |= os.O_NOFOLLOW
    try:
        descriptor = os.open(path, flags)
    except (FileNotFoundError, NotADirectoryError, PermissionError) as error:
        raise ValueError("TRITRACK_PAPER_INPUT_UNREADABLE") from error
    except OSError as error:
        raise ValueError(invalid_code) from error
    try:
        metadata = os.fstat(descriptor)
        if not stat.S_ISREG(metadata.st_mode) or not 0 < metadata.st_size <= limit:
            raise ValueError(invalid_code)
        with os.fdopen(descriptor, "rb") as stream:
            descriptor = -1
            encoded = stream.read(limit + 1)
        if len(encoded) > limit:
            raise ValueError(invalid_code)
        return encoded
    except OSError as error:
        raise ValueError(invalid_code) from error
    finally:
        if descriptor >= 0:
            os.close(descriptor)


def _load_json(
    path: Path,
    *,
    contract: str,
    invalid_code: str,
) -> LoadedArtifact:
    selected = Path(path)
    encoded = _read_regular_bytes(
        selected,
        limit=_JSON_LIMIT_BYTES,
        invalid_code=invalid_code,
    )
    try:
        payload = json.loads(encoded.decode("utf-8", errors="strict"))
        validate_contract(contract, payload)
    except (UnicodeError, json.JSONDecodeError, ValidationError) as error:
        raise ValueError(invalid_code) from error
    return LoadedArtifact(
        selected,
        payload,
        encoded,
        hashlib.sha256(encoded).hexdigest(),
        invalid_code,
        _JSON_LIMIT_BYTES,
    )


def _verify_unchanged(artifact: LoadedArtifact) -> None:
    try:
        encoded = _read_regular_bytes(
            artifact.path,
            limit=artifact.limit,
            invalid_code=artifact.invalid_code,
        )
    except ValueError as error:
        raise ValueError("TRITRACK_PAPER_INPUT_CHANGED") from error
    if hashlib.sha256(encoded).hexdigest() != artifact.sha256:
        raise ValueError("TRITRACK_PAPER_INPUT_CHANGED")


def _literal(cell: Cell, value: object, *, text_format: bool = False) -> None:
    cell.value = value
    if isinstance(value, str):
        cell.data_type = "s"
    if text_format:
        cell.number_format = "@"


def _write_row(
    worksheet,
    row: int,
    values: Sequence[object],
    *,
    text_columns: frozenset[int] = frozenset(),
) -> None:
    for column, value in enumerate(values, start=1):
        _literal(
            worksheet.cell(row=row, column=column),
            value,
            text_format=column in text_columns,
        )


def _paper_aligned_index(aligned: object) -> organizer.AlignedIndex:
    try:
        aligned_index = organizer.index_aligned_transcript(aligned)
    except ValueError as error:
        raise ValueError("TRITRACK_PAPER_ALIGNED_INVALID") from error
    if any(
        not all(
            character in "\t\n\r"
            or 0x20 <= ord(character) <= 0xD7FF
            or 0xE000 <= ord(character) <= 0xFFFD
            or 0x10000 <= ord(character) <= 0x10FFFF
            for character in take_id
        )
        for take_id in aligned_index.takes
    ):
        raise ValueError("TRITRACK_PAPER_ALIGNED_INVALID")
    return aligned_index


def _cue_rows(aligned: Mapping[str, object]) -> list[tuple[object, ...]]:
    _paper_aligned_index(aligned)
    rows: list[tuple[object, ...]] = []
    takes = aligned["takes"]
    assert isinstance(takes, list)
    for take in takes:
        assert isinstance(take, Mapping)
        if take["status"] != "completed":
            continue
        cues = take["cues"]
        assert isinstance(cues, list)
        for cue in cues:
            assert isinstance(cue, Mapping)
            rows.append(
                (
                    take["takeId"],
                    take["sourceSha256"],
                    cue["cueId"],
                    cue["startMs"],
                    cue["endMs"],
                    cue["text"],
                    cue["disposition"],
                )
            )
    return rows


def _cues_grid_sha256(rows: Sequence[Sequence[object]]) -> str:
    encoded = json.dumps(
        [list(CUES_HEADERS), *[list(row) for row in rows]],
        ensure_ascii=False,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _project_grouping(
    workbook: Workbook,
    grouping: Mapping[str, object] | None,
) -> tuple[int, int]:
    questions_sheet = workbook["Questions"]
    selections_sheet = workbook["Selections"]
    _write_row(questions_sheet, 1, QUESTIONS_HEADERS, text_columns=frozenset({1}))
    _write_row(
        selections_sheet,
        1,
        SELECTIONS_HEADERS,
        text_columns=frozenset({1, 2, 3, 5, 6, 7}),
    )
    if grouping is None:
        return 0, 0

    questions = grouping["questions"]
    reserve = grouping["reserve"]
    assert isinstance(questions, list)
    assert isinstance(reserve, list)
    question_count = 0
    selection_count = 0
    for question in sorted(questions, key=lambda item: item["order"]):
        assert isinstance(question, Mapping)
        question_count += 1
        _write_row(
            questions_sheet,
            question_count + 1,
            (question["id"], question["question"], question["order"]),
            text_columns=frozenset({1}),
        )
        answers = question["answers"]
        assert isinstance(answers, list)
        for answer in sorted(answers, key=lambda item: item["order"]):
            assert isinstance(answer, Mapping)
            selection_count += 1
            _write_row(
                selections_sheet,
                selection_count + 1,
                (
                    "ANSWER",
                    answer["id"],
                    question["id"],
                    answer["order"],
                    answer["takeId"],
                    answer["startCueId"],
                    answer["endCueId"],
                    None,
                    answer.get("note"),
                ),
                text_columns=frozenset({1, 2, 3, 5, 6, 7}),
            )
    for item in sorted(reserve, key=lambda candidate: candidate["order"]):
        assert isinstance(item, Mapping)
        selection_count += 1
        _write_row(
            selections_sheet,
            selection_count + 1,
            (
                "RESERVE",
                item["id"],
                None,
                item["order"],
                item["takeId"],
                item["startCueId"],
                item["endCueId"],
                item["reason"],
                item.get("note"),
            ),
            text_columns=frozenset({1, 2, 3, 5, 6, 7}),
        )
    return question_count, selection_count


def _build_workbook(
    aligned: Mapping[str, object],
    *,
    aligned_sha256: str,
    grouping: Mapping[str, object] | None,
) -> tuple[Workbook, dict[str, int]]:
    cue_rows = _cue_rows(aligned)
    workbook = Workbook()
    cues_sheet = workbook.active
    cues_sheet.title = "Cues"
    workbook.create_sheet("Questions")
    workbook.create_sheet("Selections")
    manifest_sheet = workbook.create_sheet("_TriTrack")
    manifest_sheet.sheet_state = "hidden"

    _write_row(cues_sheet, 1, CUES_HEADERS, text_columns=frozenset({1, 2, 3}))
    for row_number, row in enumerate(cue_rows, start=2):
        _write_row(
            cues_sheet,
            row_number,
            row,
            text_columns=frozenset({1, 2, 3}),
        )
    question_count, selection_count = _project_grouping(workbook, grouping)
    _write_row(manifest_sheet, 1, MANIFEST_HEADERS)
    manifest = (
        ("WorkbookSchemaVersion", WORKBOOK_SCHEMA_VERSION),
        ("ToolVersion", __version__),
        ("AlignedTranscriptSha256", aligned_sha256),
        ("CuesGridSha256", _cues_grid_sha256(cue_rows)),
    )
    for row_number, row in enumerate(manifest, start=2):
        _write_row(manifest_sheet, row_number, row, text_columns=frozenset({1, 2}))
    return workbook, {
        "cueCount": len(cue_rows),
        "questionCount": question_count,
        "selectionCount": selection_count,
    }


def _publish_bytes(encoded: bytes, output_path: Path) -> None:
    destination = require_absent_output(output_path)
    if not destination.parent.is_dir():
        raise ValueError("TRITRACK_OUTPUT_PARENT_MISSING")
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{destination.name}.",
        suffix=".tmp",
        dir=destination.parent,
    )
    temporary_path = Path(temporary_name)
    try:
        with os.fdopen(descriptor, "wb") as stream:
            stream.write(encoded)
            stream.flush()
            os.fsync(stream.fileno())
        try:
            os.link(temporary_path, destination)
        except FileExistsError as error:
            raise ValueError("TRITRACK_OUTPUT_EXISTS") from error
    finally:
        temporary_path.unlink(missing_ok=True)


def export_workbook(
    aligned_path: Path,
    *,
    grouping_path: Path | None,
    output_path: Path,
) -> dict[str, int]:
    """Export one strict aligned authority to an editor-facing workbook."""

    destination = require_absent_output(output_path)
    if not destination.parent.is_dir():
        raise ValueError("TRITRACK_OUTPUT_PARENT_MISSING")
    aligned = _load_json(
        aligned_path,
        contract="aligned-transcript-v1",
        invalid_code="TRITRACK_PAPER_ALIGNED_INVALID",
    )
    assert isinstance(aligned.payload, Mapping)
    grouping: LoadedArtifact | None = None
    grouping_payload: Mapping[str, object] | None = None
    aligned_index = _paper_aligned_index(aligned.payload)
    if grouping_path is not None:
        grouping = _load_json(
            grouping_path,
            contract="grouping-v1",
            invalid_code="TRITRACK_PAPER_GROUPING_INVALID",
        )
        if grouping.encoded != organizer.encode_grouping(grouping.payload):
            raise ValueError("TRITRACK_PAPER_GROUPING_INVALID")
        grouping_payload = organizer.validate_grouping(
            grouping.payload,
            aligned_index=aligned_index,
            aligned_sha256=aligned.sha256,
        )

    workbook, summary = _build_workbook(
        aligned.payload,
        aligned_sha256=aligned.sha256,
        grouping=grouping_payload,
    )
    buffer = io.BytesIO()
    workbook.save(buffer)
    _verify_unchanged(aligned)
    if grouping is not None:
        _verify_unchanged(grouping)
    _publish_bytes(buffer.getvalue(), destination)
    return summary


def _load_workbook_artifact(path: Path) -> tuple[LoadedArtifact, Workbook]:
    selected = Path(path)
    encoded = _read_regular_bytes(
        selected,
        limit=_WORKBOOK_LIMIT_BYTES,
        invalid_code="TRITRACK_PAPER_WORKBOOK_INVALID",
    )
    artifact = LoadedArtifact(
        selected,
        None,
        encoded,
        hashlib.sha256(encoded).hexdigest(),
        "TRITRACK_PAPER_WORKBOOK_INVALID",
        _WORKBOOK_LIMIT_BYTES,
    )
    try:
        with zipfile.ZipFile(io.BytesIO(encoded)) as archive:
            members = archive.infolist()
            names = [member.filename for member in members]
            expanded_size = 0
            if (
                not members
                or len(members) > _WORKBOOK_MEMBER_LIMIT
                or len(names) != len(set(names))
            ):
                raise ValueError("TRITRACK_PAPER_WORKBOOK_INVALID")
            for member in members:
                member_path = PurePosixPath(member.filename)
                if (
                    member.flag_bits & 0x1
                    or member.filename.startswith(("/", "\\"))
                    or "\\" in member.filename
                    or ".." in member_path.parts
                    or member.filename.lower().endswith("vbaproject.bin")
                    or member.file_size > _WORKBOOK_SINGLE_MEMBER_LIMIT_BYTES
                ):
                    raise ValueError("TRITRACK_PAPER_WORKBOOK_INVALID")
                expanded_size += member.file_size
                if expanded_size > _WORKBOOK_EXPANDED_LIMIT_BYTES:
                    raise ValueError("TRITRACK_PAPER_WORKBOOK_INVALID")
    except ValueError:
        raise
    except (
        OSError,
        KeyError,
        TypeError,
        zipfile.BadZipFile,
        InvalidFileException,
        element_tree.ParseError,
    ) as error:
        raise ValueError("TRITRACK_PAPER_WORKBOOK_INVALID") from error

    try:
        workbook = load_workbook(
            io.BytesIO(encoded),
            data_only=False,
            read_only=False,
            keep_links=True,
        )
    except (
        OSError,
        KeyError,
        TypeError,
        ValueError,
        zipfile.BadZipFile,
        InvalidFileException,
        element_tree.ParseError,
    ) as error:
        raise ValueError("TRITRACK_PAPER_WORKBOOK_INVALID") from error
    return artifact, workbook


def _reject_unsafe_workbook_state(
    workbook: Workbook,
    *,
    cue_row_count: int,
) -> None:
    if workbook.sheetnames != list(SHEET_NAMES):
        raise ValueError("TRITRACK_PAPER_SHEETS_INVALID")
    if workbook["_TriTrack"].sheet_state != "hidden" or any(
        workbook[name].sheet_state != "visible" for name in SHEET_NAMES[:-1]
    ):
        raise ValueError("TRITRACK_PAPER_WORKBOOK_INVALID")
    if len(workbook.defined_names) or getattr(workbook, "_external_links", []):
        raise ValueError("TRITRACK_PAPER_WORKBOOK_INVALID")
    maximum_dimensions = {
        "Cues": (cue_row_count + 1, len(CUES_HEADERS)),
        "Questions": (cue_row_count + 1, len(QUESTIONS_HEADERS)),
        "Selections": (cue_row_count + 1, len(SELECTIONS_HEADERS)),
        "_TriTrack": (5, len(MANIFEST_HEADERS)),
    }
    for worksheet in workbook.worksheets:
        maximum_rows, maximum_columns = maximum_dimensions[worksheet.title]
        if (
            worksheet.max_row > maximum_rows
            or worksheet.max_column > maximum_columns
        ):
            raise ValueError("TRITRACK_PAPER_WORKBOOK_INVALID")
    for worksheet in workbook.worksheets:
        if worksheet.merged_cells.ranges:
            raise ValueError("TRITRACK_PAPER_WORKBOOK_INVALID")
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.hyperlink is not None:
                    raise ValueError("TRITRACK_PAPER_WORKBOOK_INVALID")
                if cell.data_type == "f":
                    raise ValueError("TRITRACK_PAPER_FORMULA_FORBIDDEN")


def _sheet_rows(
    worksheet,
    headers: Sequence[str],
    *,
    invalid_code: str,
) -> list[tuple[object, ...]]:
    actual_headers = tuple(
        worksheet.cell(row=1, column=column).value
        for column in range(1, len(headers) + 1)
    )
    if actual_headers != tuple(headers):
        raise ValueError(invalid_code)
    if any(
        worksheet.cell(row=row, column=column).value is not None
        for row in range(1, worksheet.max_row + 1)
        for column in range(len(headers) + 1, worksheet.max_column + 1)
    ):
        raise ValueError(invalid_code)
    raw_rows = [
        tuple(
            worksheet.cell(row=row, column=column).value
            for column in range(1, len(headers) + 1)
        )
        for row in range(2, worksheet.max_row + 1)
    ]
    while raw_rows and all(value is None for value in raw_rows[-1]):
        raw_rows.pop()
    if any(all(value is None for value in row) for row in raw_rows):
        raise ValueError(invalid_code)
    return raw_rows


def _require_exact_row_types(
    actual: Sequence[object],
    expected: Sequence[object],
) -> None:
    if len(actual) != len(expected) or any(
        value != reference or type(value) is not type(reference)
        for value, reference in zip(actual, expected, strict=True)
    ):
        raise ValueError("TRITRACK_PAPER_REFERENCE_MISMATCH")


def _verify_cues_grid(
    workbook: Workbook,
    aligned: Mapping[str, object],
) -> list[tuple[object, ...]]:
    expected = _cue_rows(aligned)
    actual = _sheet_rows(
        workbook["Cues"],
        CUES_HEADERS,
        invalid_code="TRITRACK_PAPER_REFERENCE_MISMATCH",
    )
    if len(actual) != len(expected):
        raise ValueError("TRITRACK_PAPER_REFERENCE_MISMATCH")
    for actual_row, expected_row in zip(actual, expected, strict=True):
        _require_exact_row_types(actual_row, expected_row)
    return expected


def _verify_manifest(
    workbook: Workbook,
    *,
    aligned_sha256: str,
    cue_rows: Sequence[Sequence[object]],
) -> None:
    expected = [
        ("WorkbookSchemaVersion", WORKBOOK_SCHEMA_VERSION),
        ("ToolVersion", __version__),
        ("AlignedTranscriptSha256", aligned_sha256),
        ("CuesGridSha256", _cues_grid_sha256(cue_rows)),
    ]
    actual = _sheet_rows(
        workbook["_TriTrack"],
        MANIFEST_HEADERS,
        invalid_code="TRITRACK_PAPER_MANIFEST_MISMATCH",
    )
    if actual != expected:
        raise ValueError("TRITRACK_PAPER_MANIFEST_MISMATCH")


def _canonical_workbook_text(
    value: object,
    *,
    maximum: int,
    required: bool,
) -> str | None:
    try:
        return organizer.canonical_editor_text(
            value,
            maximum=maximum,
            required=required,
        )
    except (TypeError, ValueError) as error:
        raise ValueError("TRITRACK_PAPER_ROW_INVALID") from error


def _positive_integer(value: object) -> int:
    if type(value) is not int or value < 1:
        raise ValueError("TRITRACK_PAPER_ROW_INVALID")
    return value


def _required_string(value: object) -> str:
    if not isinstance(value, str) or not value:
        raise ValueError("TRITRACK_PAPER_ROW_INVALID")
    return value


def _grouping_from_workbook(
    workbook: Workbook,
    *,
    aligned_index: organizer.AlignedIndex,
    aligned_sha256: str,
) -> dict[str, object]:
    question_rows = _sheet_rows(
        workbook["Questions"],
        QUESTIONS_HEADERS,
        invalid_code="TRITRACK_PAPER_ROW_INVALID",
    )
    selection_rows = _sheet_rows(
        workbook["Selections"],
        SELECTIONS_HEADERS,
        invalid_code="TRITRACK_PAPER_ROW_INVALID",
    )
    questions: list[dict[str, object]] = []
    questions_by_id: dict[str, dict[str, object]] = {}
    for question_id, question_text, order in question_rows:
        identifier = _required_string(question_id)
        question = {
            "id": identifier,
            "question": _canonical_workbook_text(
                question_text,
                maximum=organizer.QUESTION_TEXT_LIMIT,
                required=True,
            ),
            "order": _positive_integer(order),
            "answers": [],
        }
        if identifier in questions_by_id:
            raise ValueError("TRITRACK_ORGANIZER_DUPLICATE_ID")
        questions.append(question)
        questions_by_id[identifier] = question

    reserve: list[dict[str, object]] = []
    for row in selection_rows:
        (
            placement,
            segment_id,
            question_id,
            order,
            take_id,
            start_cue_id,
            end_cue_id,
            reserve_reason,
            editor_note,
        ) = row
        placement = _required_string(placement)
        common: dict[str, object] = {
            "id": _required_string(segment_id),
            "order": _positive_integer(order),
            "takeId": _required_string(take_id),
            "startCueId": _required_string(start_cue_id),
            "endCueId": _required_string(end_cue_id),
        }
        note = _canonical_workbook_text(
            editor_note,
            maximum=organizer.NOTE_TEXT_LIMIT,
            required=False,
        )
        if note is not None:
            common["note"] = note
        if placement == "ANSWER":
            if reserve_reason not in {None, ""}:
                raise ValueError("TRITRACK_PAPER_ROW_INVALID")
            selected_question = questions_by_id.get(_required_string(question_id))
            if selected_question is None:
                raise ValueError("TRITRACK_PAPER_ROW_INVALID")
            answers = selected_question["answers"]
            assert isinstance(answers, list)
            answers.append(common)
        elif placement == "RESERVE":
            if question_id not in {None, ""}:
                raise ValueError("TRITRACK_PAPER_ROW_INVALID")
            common["reason"] = _canonical_workbook_text(
                reserve_reason,
                maximum=organizer.QUESTION_TEXT_LIMIT,
                required=True,
            )
            reserve.append(common)
        else:
            raise ValueError("TRITRACK_PAPER_ROW_INVALID")

    grouping: dict[str, object] = {
        "schemaVersion": "tritrack.grouping/v1",
        "alignedTranscriptSha256": aligned_sha256,
        "questions": questions,
        "reserve": reserve,
    }
    return organizer.validate_grouping(
        grouping,
        aligned_index=aligned_index,
        aligned_sha256=aligned_sha256,
    )


def validate_workbook(
    aligned_path: Path,
    workbook_path: Path,
) -> ValidatedWorkbook:
    """Validate and re-derive one workbook without publishing output."""

    aligned = _load_json(
        aligned_path,
        contract="aligned-transcript-v1",
        invalid_code="TRITRACK_PAPER_ALIGNED_INVALID",
    )
    workbook_artifact, workbook = _load_workbook_artifact(workbook_path)
    assert isinstance(aligned.payload, Mapping)
    aligned_index = _paper_aligned_index(aligned.payload)
    cue_rows = _cue_rows(aligned.payload)
    _reject_unsafe_workbook_state(workbook, cue_row_count=len(cue_rows))
    cue_rows = _verify_cues_grid(workbook, aligned.payload)
    _verify_manifest(
        workbook,
        aligned_sha256=aligned.sha256,
        cue_rows=cue_rows,
    )
    grouping = _grouping_from_workbook(
        workbook,
        aligned_index=aligned_index,
        aligned_sha256=aligned.sha256,
    )
    _verify_unchanged(aligned)
    _verify_unchanged(workbook_artifact)
    questions = grouping["questions"]
    reserve = grouping["reserve"]
    assert isinstance(questions, list)
    assert isinstance(reserve, list)
    answer_count = 0
    for question in questions:
        assert isinstance(question, Mapping)
        answers = question["answers"]
        assert isinstance(answers, list)
        answer_count += len(answers)
    return ValidatedWorkbook(
        aligned_sha256=aligned.sha256,
        workbook_sha256=workbook_artifact.sha256,
        workbook_schema_version=WORKBOOK_SCHEMA_VERSION,
        cue_count=len(cue_rows),
        question_count=len(questions),
        answer_count=answer_count,
        reserve_count=len(reserve),
        grouping=grouping,
    )


def apply_workbook(
    aligned_path: Path,
    workbook_path: Path,
    *,
    output_path: Path,
) -> dict[str, object]:
    """Apply strict workbook intent and publish canonical grouping JSON."""

    destination = require_absent_output(output_path)
    if not destination.parent.is_dir():
        raise ValueError("TRITRACK_OUTPUT_PARENT_MISSING")
    validated = validate_workbook(aligned_path, workbook_path)
    _publish_bytes(organizer.encode_grouping(validated.grouping), destination)
    return validated.grouping
